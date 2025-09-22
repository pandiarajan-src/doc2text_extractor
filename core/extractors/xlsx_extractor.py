import io
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

from .base import BaseExtractor, ExtractionResult

logger = logging.getLogger(__name__)


class XLSXExtractor(BaseExtractor):
    def __init__(self):
        super().__init__()
        self.supported_extensions = [".xlsx", ".xls"]
        self.supported_mime_types = [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ]

    def extract(self, file_path: Path, output_dir: Path) -> ExtractionResult:
        try:
            images_dir = output_dir / "images"
            images_dir.mkdir(exist_ok=True)

            workbook = load_workbook(file_path, data_only=True)
            extracted_images = []
            text_content = []

            metadata = self.get_file_metadata(file_path)

            props = workbook.properties
            if props:
                metadata.title = props.title
                metadata.author = props.creator
                metadata.subject = props.subject

                if props.keywords:
                    metadata.keywords = [k.strip() for k in props.keywords.split(",")]

                metadata.document_properties = {
                    "category": props.category,
                    "comments": props.description,
                    "company": getattr(props, "company", None),
                    "manager": getattr(props, "manager", None),
                    "created": props.created.isoformat() if props.created else None,
                    "modified": props.modified.isoformat() if props.modified else None,
                    "last_modified_by": props.lastModifiedBy,
                    "revision": getattr(props, "revision", None),
                    "version": getattr(props, "version", None),
                    "sheets_count": len(workbook.worksheets),
                    "sheet_names": [ws.title for ws in workbook.worksheets],
                }

            img_counter = 1

            for sheet_index, worksheet in enumerate(workbook.worksheets):
                text_content.append(f"=== Sheet: {worksheet.title} ===\n")

                if hasattr(worksheet, "_images") and worksheet._images:
                    for image in worksheet._images:
                        try:
                            img_data = image.ref
                            if hasattr(img_data, "blob"):
                                image_filename = (
                                    f"xlsx_sheet_{sheet_index + 1}_img_{img_counter}.png"
                                )
                                image_path = images_dir / image_filename

                                pil_image = Image.open(io.BytesIO(img_data.blob))
                                pil_image.save(image_path, "PNG")

                                extracted_images.append(str(image_path))
                                img_counter += 1
                        except Exception as e:
                            logger.warning(
                                f"Failed to extract image from sheet {worksheet.title}: {e}"
                            )
                            continue

                max_row = worksheet.max_row
                max_col = worksheet.max_column

                if max_row and max_col and max_row > 0 and max_col > 0:
                    for row_num in range(1, min(max_row + 1, 10000)):
                        row_data = []
                        has_data = False

                        for col_num in range(1, min(max_col + 1, 100)):
                            cell = worksheet.cell(row=row_num, column=col_num)
                            cell_value = cell.value

                            if cell_value is not None:
                                has_data = True
                                if isinstance(cell_value, datetime):
                                    cell_value = cell_value.isoformat()
                                else:
                                    cell_value = str(cell_value)
                            else:
                                cell_value = ""

                            row_data.append(cell_value)

                        if has_data:
                            text_content.append(" | ".join(row_data))
                        elif row_num > 100 and not has_data:
                            break

                text_content.append("\n")

            full_text = "\n".join(text_content)

            content_file = output_dir / "content.txt"
            with open(content_file, "w", encoding="utf-8") as f:
                f.write(full_text)

            self.save_metadata(metadata, output_dir)

            return ExtractionResult(
                text=full_text, images=extracted_images, metadata=metadata.to_dict(), success=True
            )

        except Exception as e:
            logger.error(f"XLSX extraction failed for {file_path}: {e}")
            return ExtractionResult(text="", images=[], metadata={}, success=False, error=str(e))
